import xlsxwriter
from openpyxl import load_workbook
import datetime

def save_to_excel(billnumber,name, phone, bathsoap,bodylotion,facewash,facecream,hairspray,hairgel,
                  rice,oil,daal,Wheat,Sugar,Tea,
                  Maaza,Pepsi,Sprite,Dew,Frooti,Cococola):

    # Try to load the existing workbook
    try:
        workbook = load_workbook('billing_System.xlsx')
        worksheet = workbook.active
    except FileNotFoundError:
        # If the file does not exist, create a new workbook and worksheet
        workbook = xlsxwriter.Workbook('billing_System.xlsx')
        worksheet = workbook.add_worksheet()

        # Add Header
        worksheet.write('A1','Bill Number')
        worksheet.write('B1','Name')
        worksheet.write('C1','Phone Number')
        worksheet.write('D1','Date')

        worksheet.write('E1','Bath Soap')
        worksheet.write('F1','Body Lotion')
        worksheet.write('G1','Face Wash')
        worksheet.write('H1','Face Cream')
        worksheet.write('I1','Hair Spray')
        worksheet.write('J1', 'Hair Gel')

        worksheet.write('K1','Rice')
        worksheet.write('L1','Oil')
        worksheet.write('M1','Daal')
        worksheet.write('N1','Wheat')
        worksheet.write('O1','Suagr')
        worksheet.write('P1','Tea')

        worksheet.write('Q1','Maaza')
        worksheet.write('R1','Pepsi')
        worksheet.write('S1','Sprite')
        worksheet.write('T1','Dew')
        worksheet.write('U1','Frooti')
        worksheet.write('V1','Coco Cola')

        # Save and close the workbook
        workbook.close()

        # Reopen the workbook with openpyxl to append data
        workbook = load_workbook('billing_System.xlsx')
        worksheet = workbook.active

    # Find the next available row
    next_row = worksheet.max_row + 1

    # Add Data
    worksheet.cell(row=next_row,column=1,value=billnumber)
    worksheet.cell(row=next_row, column=2, value=name)
    worksheet.cell(row=next_row, column=3, value=phone)
    worksheet.cell(row=next_row,column=4,value=datetime.datetime.now())

    worksheet.cell(row=next_row, column=5, value=bathsoap)
    worksheet.cell(row=next_row, column=6, value=facewash)
    worksheet.cell(row=next_row, column=7, value=bodylotion)
    worksheet.cell(row=next_row, column=8, value=facecream)
    worksheet.cell(row=next_row, column=9, value=hairspray)
    worksheet.cell(row=next_row, column=10, value=hairgel)

    worksheet.cell(row=next_row, column=11, value=rice)
    worksheet.cell(row=next_row, column=12, value=oil)
    worksheet.cell(row=next_row, column=13, value=daal)
    worksheet.cell(row=next_row, column=14, value=Wheat)
    worksheet.cell(row=next_row, column=15, value=Sugar)
    worksheet.cell(row=next_row, column=16, value=Tea)

    worksheet.cell(row=next_row, column=17, value=Maaza)
    worksheet.cell(row=next_row, column=18, value=Pepsi)
    worksheet.cell(row=next_row, column=19, value=Sprite)
    worksheet.cell(row=next_row, column=20, value=Dew)
    worksheet.cell(row=next_row, column=21, value=Frooti)
    worksheet.cell(row=next_row, column=22, value=Cococola)

    # Save the workbook
    workbook.save('billing_System.xlsx')



